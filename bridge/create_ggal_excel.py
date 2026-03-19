"""
Genera Excel con opciones reales de GGAL + acción y futuro Rofex.
Símbolos según nomenclatura BYMA/etrader.

Uso (en Windows PowerShell):
    python create_ggal_excel.py
"""
try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date
import math
import random

wb = Workbook()

# Estilos
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
call_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
put_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
stock_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
future_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# ============================================================
# Datos de referencia GGAL
# ============================================================
GGAL_SPOT = 637.47  # Precio acción GGAL (pesos argentinos, mercado local)

# Vencimientos
# A = Abril 2026, J = Junio 2026
VENCIMIENTOS = {
    "A": date(2026, 4, 17),   # Tercer viernes de Abril
    "J": date(2026, 6, 19),   # Tercer viernes de Junio
}

# Parsear símbolos reales
# GFGC43747A = GFG (GGAL) C (Call) 43747 (strike en centavos/base) A (Abril)
# El strike se lee como: 437.47, 457.47, etc.

OPCIONES_ABRIL = [
    # Calls Abril
    ("GFGC43747A", "CALL", 437.47), ("GFGC45747A", "CALL", 457.47),
    ("GFGC47747A", "CALL", 477.47), ("GFGC49747A", "CALL", 497.47),
    ("GFGC51747A", "CALL", 517.47), ("GFGC53747A", "CALL", 537.47),
    ("GFGC55747A", "CALL", 557.47), ("GFGC57747A", "CALL", 577.47),
    ("GFGC59501A", "CALL", 595.01), ("GFGC61262A", "CALL", 612.62),
    ("GFGC63747A", "CALL", 637.47), ("GFGC65747A", "CALL", 657.47),
    ("GFGC67747A", "CALL", 677.47), ("GFGC69029A", "CALL", 690.29),
    ("GFGC71747A", "CALL", 717.47), ("GFGC73262A", "CALL", 732.62),
    ("GFGC75029A", "CALL", 750.29), ("GFGC77262A", "CALL", 772.62),
    ("GFGC79262A", "CALL", 792.62), ("GFGC82029A", "CALL", 820.29),
    ("GFGC85501A", "CALL", 855.01), ("GFGC88262A", "CALL", 882.62),
    ("GFGC91501A", "CALL", 915.01), ("GFGC94501A", "CALL", 945.01),
    ("GFGC96801A", "CALL", 968.01), ("GFGC10126A", "CALL", 1012.60),
    ("GFGC10950A", "CALL", 1095.00), ("GFGC11350A", "CALL", 1135.00),
    ("GFGC11775A", "CALL", 1177.50),
    # Puts Abril
    ("GFGV43747A", "PUT", 437.47), ("GFGV45747A", "PUT", 457.47),
    ("GFGV47747A", "PUT", 477.47), ("GFGV49747A", "PUT", 497.47),
    ("GFGV51747A", "PUT", 517.47), ("GFGV53747A", "PUT", 537.47),
    ("GFGV55747A", "PUT", 557.47), ("GFGV57747A", "PUT", 577.47),
    ("GFGV59501A", "PUT", 595.01), ("GFGV61262A", "PUT", 612.62),
    ("GFGV63747A", "PUT", 637.47), ("GFGV65747A", "PUT", 657.47),
    ("GFGV67747A", "PUT", 677.47), ("GFGV69029A", "PUT", 690.29),
    ("GFGV71747A", "PUT", 717.47), ("GFGV73262A", "PUT", 732.62),
    ("GFGV75029A", "PUT", 750.29), ("GFGV77262A", "PUT", 772.62),
    ("GFGV79262A", "PUT", 792.62), ("GFGV82029A", "PUT", 820.29),
    ("GFGV85501A", "PUT", 855.01), ("GFGV88262A", "PUT", 882.62),
    ("GFGV91501A", "PUT", 915.01), ("GFGV94501A", "PUT", 945.01),
    ("GFGV96801A", "PUT", 968.01), ("GFGV10126A", "PUT", 1012.60),
    ("GFGV10950A", "PUT", 1095.00), ("GFGV11350A", "PUT", 1135.00),
    ("GFGV11775A", "PUT", 1177.50),
]

OPCIONES_JUNIO = [
    # Calls Junio
    ("GFGC43747J", "CALL", 437.47), ("GFGC55747J", "CALL", 557.47),
    ("GFGC61747J", "CALL", 617.47), ("GFGC63747J", "CALL", 637.47),
    ("GFGC65747J", "CALL", 657.47), ("GFGC67747J", "CALL", 677.47),
    ("GFGC71747J", "CALL", 717.47), ("GFGC73747J", "CALL", 737.47),
    ("GFGC75501J", "CALL", 755.01), ("GFGC77747J", "CALL", 777.47),
    ("GFGC79747J", "CALL", 797.47), ("GFGC82747J", "CALL", 827.47),
    ("GFGC85501J", "CALL", 855.01), ("GFGC88501J", "CALL", 885.01),
    ("GFGC91501J", "CALL", 915.01), ("GFGC10550J", "CALL", 1055.00),
    ("GFGC10950J", "CALL", 1095.00), ("GFGC11350J", "CALL", 1135.00),
    # Puts Junio
    ("GFGV43747J", "PUT", 437.47), ("GFGV55747J", "PUT", 557.47),
    ("GFGV61747J", "PUT", 617.47), ("GFGV63747J", "PUT", 637.47),
    ("GFGV65747J", "PUT", 657.47), ("GFGV67747J", "PUT", 677.47),
    ("GFGV71747J", "PUT", 717.47), ("GFGV73747J", "PUT", 737.47),
    ("GFGV75501J", "PUT", 755.01), ("GFGV77747J", "PUT", 777.47),
    ("GFGV79747J", "PUT", 797.47), ("GFGV82747J", "PUT", 827.47),
    ("GFGV85501J", "PUT", 855.01), ("GFGV88501J", "PUT", 885.01),
    ("GFGV91501J", "PUT", 915.01), ("GFGV10550J", "PUT", 1055.00),
    ("GFGV10950J", "PUT", 1095.00), ("GFGV11350J", "PUT", 1135.00),
]

ALL_OPTIONS = OPCIONES_ABRIL + OPCIONES_JUNIO


def black_scholes_price(S, K, T, r, sigma, option_type):
    """Precio Black-Scholes simplificado."""
    if T <= 0 or sigma <= 0:
        return max(0, S - K) if option_type == "CALL" else max(0, K - S)

    d1 = (math.log(S / K) + (r + sigma**2 / 2) * T) / (sigma * math.sqrt(T))
    d2 = d1 - sigma * math.sqrt(T)

    # Aproximación de N(x) con función logística
    def norm_cdf(x):
        return 1 / (1 + math.exp(-1.7 * x - 0.73 * x**3 / (1 + abs(x))))

    if option_type == "CALL":
        price = S * norm_cdf(d1) - K * math.exp(-r * T) * norm_cdf(d2)
    else:
        price = K * math.exp(-r * T) * norm_cdf(-d2) - S * norm_cdf(-d1)

    return max(0.01, price)


def calc_greeks(S, K, T, r, sigma, option_type):
    """Greeks simplificados."""
    if T <= 0 or sigma <= 0:
        return {"delta": 1.0 if option_type == "CALL" else -1.0, "gamma": 0, "theta": 0, "vega": 0}

    d1 = (math.log(S / K) + (r + sigma**2 / 2) * T) / (sigma * math.sqrt(T))

    def norm_cdf(x):
        return 1 / (1 + math.exp(-1.7 * x - 0.73 * x**3 / (1 + abs(x))))

    def norm_pdf(x):
        return math.exp(-x**2 / 2) / math.sqrt(2 * math.pi)

    delta = norm_cdf(d1) if option_type == "CALL" else norm_cdf(d1) - 1
    gamma = norm_pdf(d1) / (S * sigma * math.sqrt(T))
    vega = S * norm_pdf(d1) * math.sqrt(T) / 100  # por 1% de vol
    theta = -(S * norm_pdf(d1) * sigma) / (2 * math.sqrt(T)) / 365  # por día

    return {"delta": round(delta, 4), "gamma": round(gamma, 6), "theta": round(theta, 4), "vega": round(vega, 4)}


# ============================================================
# HOJA 1: Opciones
# ============================================================
ws = wb.active
ws.title = "Opciones"

headers = [
    "Simbolo", "Subyacente", "Tipo", "Strike", "Vencimiento",
    "Ultimo", "Bid", "Ask", "Volumen", "VI",
    "OI", "Spot", "Delta", "Gamma", "Theta", "Vega"
]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Parámetros de mercado
r = 0.40  # Tasa libre de riesgo Argentina (~40% anual)
base_vol = 0.55  # Volatilidad base GGAL

row = 2
for symbol, opt_type, strike in ALL_OPTIONS:
    # Determinar vencimiento
    venc_code = symbol[-1]
    venc_date = VENCIMIENTOS.get(venc_code, date(2026, 4, 17))
    T = max(0.001, (venc_date - date.today()).days / 365.0)

    # Volatility smile
    moneyness = (GGAL_SPOT - strike) / GGAL_SPOT
    iv = base_vol + abs(moneyness) * 0.12 + random.uniform(-0.02, 0.02)
    iv = round(max(0.20, min(1.20, iv)), 4)

    # Precio BS
    price = black_scholes_price(GGAL_SPOT, strike, T, r, iv, opt_type)
    price = round(price, 2)

    # Spread bid-ask (más ancho para OTM)
    spread_pct = 0.02 + abs(moneyness) * 0.03
    bid = round(price * (1 - spread_pct), 2)
    ask = round(price * (1 + spread_pct), 2)
    bid = max(0.01, bid)
    ask = max(bid + 0.01, ask)

    # Volumen y OI (más alto cerca del ATM)
    atm_factor = max(0.1, 1 - abs(moneyness) * 3)
    vol = int(random.randint(5, 2000) * atm_factor)
    oi = int(random.randint(100, 30000) * atm_factor)

    # Greeks
    greeks = calc_greeks(GGAL_SPOT, strike, T, r, iv, opt_type)

    fill = call_fill if opt_type == "CALL" else put_fill

    ws.cell(row=row, column=1, value=symbol).border = thin_border
    ws.cell(row=row, column=2, value="GGAL").border = thin_border
    ws.cell(row=row, column=3, value=opt_type).border = thin_border
    ws.cell(row=row, column=4, value=strike).border = thin_border
    c = ws.cell(row=row, column=5, value=venc_date)
    c.border = thin_border
    c.number_format = 'DD/MM/YYYY'
    ws.cell(row=row, column=6, value=price).border = thin_border
    ws.cell(row=row, column=7, value=bid).border = thin_border
    ws.cell(row=row, column=8, value=ask).border = thin_border
    ws.cell(row=row, column=9, value=vol).border = thin_border
    c = ws.cell(row=row, column=10, value=iv)
    c.border = thin_border
    c.number_format = '0.00%'
    ws.cell(row=row, column=11, value=oi).border = thin_border
    ws.cell(row=row, column=12, value=GGAL_SPOT).border = thin_border
    ws.cell(row=row, column=13, value=greeks["delta"]).border = thin_border
    ws.cell(row=row, column=14, value=greeks["gamma"]).border = thin_border
    ws.cell(row=row, column=15, value=greeks["theta"]).border = thin_border
    ws.cell(row=row, column=16, value=greeks["vega"]).border = thin_border

    for c in range(1, 17):
        ws.cell(row=row, column=c).fill = fill

    row += 1

total_options = row - 2

# ============================================================
# HOJA 2: Subyacentes (GGAL acción + Futuro Rofex)
# ============================================================
ws_sub = wb.create_sheet("Subyacentes")

sub_headers = [
    "Simbolo", "Nombre", "Tipo", "Ultimo", "Variacion %",
    "Bid", "Ask", "Volumen", "Apertura", "Maximo", "Minimo", "Cierre Ant"
]
for col, h in enumerate(sub_headers, 1):
    cell = ws_sub.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# GGAL Acción
ggal_data = {
    "symbol": "GGAL", "name": "Grupo Financiero Galicia S.A.",
    "type": "ACCION", "last": GGAL_SPOT, "var": round(random.uniform(-2, 3), 2),
    "bid": round(GGAL_SPOT - 0.50, 2), "ask": round(GGAL_SPOT + 0.50, 2),
    "vol": random.randint(500000, 3000000),
    "open": round(GGAL_SPOT * (1 + random.uniform(-0.01, 0.01)), 2),
    "high": round(GGAL_SPOT * 1.02, 2), "low": round(GGAL_SPOT * 0.98, 2),
    "prev_close": round(GGAL_SPOT * (1 - random.uniform(0, 0.02)), 2),
}

# Futuro GGAL Rofex Abril
ggal_fut_apr = {
    "symbol": "GGAL.AB26", "name": "Futuro GGAL Abril 2026 (Rofex/MAV)",
    "type": "FUTURO", "last": round(GGAL_SPOT * 1.035, 2),
    "var": round(random.uniform(-2, 3), 2),
    "bid": round(GGAL_SPOT * 1.033, 2), "ask": round(GGAL_SPOT * 1.037, 2),
    "vol": random.randint(10000, 100000),
    "open": round(GGAL_SPOT * 1.032, 2),
    "high": round(GGAL_SPOT * 1.045, 2), "low": round(GGAL_SPOT * 1.025, 2),
    "prev_close": round(GGAL_SPOT * 1.030, 2),
}

# Futuro GGAL Rofex Junio
ggal_fut_jun = {
    "symbol": "GGAL.JN26", "name": "Futuro GGAL Junio 2026 (Rofex/MAV)",
    "type": "FUTURO", "last": round(GGAL_SPOT * 1.075, 2),
    "var": round(random.uniform(-2, 3), 2),
    "bid": round(GGAL_SPOT * 1.073, 2), "ask": round(GGAL_SPOT * 1.077, 2),
    "vol": random.randint(5000, 50000),
    "open": round(GGAL_SPOT * 1.072, 2),
    "high": round(GGAL_SPOT * 1.085, 2), "low": round(GGAL_SPOT * 1.065, 2),
    "prev_close": round(GGAL_SPOT * 1.070, 2),
}

for i, data in enumerate([ggal_data, ggal_fut_apr, ggal_fut_jun], 2):
    fill = stock_fill if data["type"] == "ACCION" else future_fill
    ws_sub.cell(row=i, column=1, value=data["symbol"]).border = thin_border
    ws_sub.cell(row=i, column=2, value=data["name"]).border = thin_border
    ws_sub.cell(row=i, column=3, value=data["type"]).border = thin_border
    ws_sub.cell(row=i, column=4, value=data["last"]).border = thin_border
    c = ws_sub.cell(row=i, column=5, value=data["var"])
    c.border = thin_border
    c.number_format = '0.00%'
    c.font = Font(color="006400" if data["var"] >= 0 else "8B0000", bold=True)
    ws_sub.cell(row=i, column=6, value=data["bid"]).border = thin_border
    ws_sub.cell(row=i, column=7, value=data["ask"]).border = thin_border
    ws_sub.cell(row=i, column=8, value=data["vol"]).border = thin_border
    ws_sub.cell(row=i, column=9, value=data["open"]).border = thin_border
    ws_sub.cell(row=i, column=10, value=data["high"]).border = thin_border
    ws_sub.cell(row=i, column=11, value=data["low"]).border = thin_border
    ws_sub.cell(row=i, column=12, value=data["prev_close"]).border = thin_border
    for c in range(1, 13):
        ws_sub.cell(row=i, column=c).fill = fill

# Anchos
for i, w in enumerate([16, 40, 10, 10, 12, 10, 10, 12, 10, 10, 10, 12], 1):
    ws_sub.column_dimensions[chr(64 + i)].width = w

# ============================================================
# HOJA 3: Portfolio (posiciones de ejemplo)
# ============================================================
ws_port = wb.create_sheet("Portfolio")

port_headers = [
    "Simbolo", "Cantidad", "Precio Entrada", "Precio Actual",
    "P&L", "Tipo", "Subyacente", "Strike", "Vencimiento"
]
for col, h in enumerate(port_headers, 1):
    cell = ws_port.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Ejemplo: Bull Call Spread + Protective Put
posiciones = [
    # Bull Call Spread Abril: compra 637.47 call, vende 717.47 call
    {"sym": "GFGC63747A", "qty": 100, "entry": 45.0, "type": "CALL", "strike": 637.47, "venc": date(2026, 4, 17)},
    {"sym": "GFGC71747A", "qty": -100, "entry": 12.0, "type": "CALL", "strike": 717.47, "venc": date(2026, 4, 17)},
    # Protective Put
    {"sym": "GFGV59501A", "qty": 50, "entry": 8.0, "type": "PUT", "strike": 595.01, "venc": date(2026, 4, 17)},
    # Straddle Junio vendido
    {"sym": "GFGC63747J", "qty": -20, "entry": 85.0, "type": "CALL", "strike": 637.47, "venc": date(2026, 6, 19)},
    {"sym": "GFGV63747J", "qty": -20, "entry": 65.0, "type": "PUT", "strike": 637.47, "venc": date(2026, 6, 19)},
    # Acción GGAL
    {"sym": "GGAL", "qty": 500, "entry": 620.0, "type": "ACCION", "strike": 0, "venc": None},
]

for i, pos in enumerate(posiciones, 2):
    # Calcular precio actual
    if pos["type"] == "ACCION":
        current = GGAL_SPOT
    else:
        T = max(0.001, (pos["venc"] - date.today()).days / 365.0) if pos["venc"] else 0.001
        current = round(black_scholes_price(GGAL_SPOT, pos["strike"], T, r, base_vol, pos["type"]), 2)

    pnl = round((current - pos["entry"]) * pos["qty"], 2)

    ws_port.cell(row=i, column=1, value=pos["sym"]).border = thin_border
    ws_port.cell(row=i, column=2, value=pos["qty"]).border = thin_border
    ws_port.cell(row=i, column=3, value=pos["entry"]).border = thin_border
    ws_port.cell(row=i, column=4, value=current).border = thin_border
    c = ws_port.cell(row=i, column=5, value=pnl)
    c.border = thin_border
    c.font = Font(color="006400" if pnl >= 0 else "8B0000", bold=True)
    ws_port.cell(row=i, column=6, value=pos["type"]).border = thin_border
    ws_port.cell(row=i, column=7, value="GGAL").border = thin_border
    ws_port.cell(row=i, column=8, value=pos["strike"]).border = thin_border
    if pos["venc"]:
        c = ws_port.cell(row=i, column=9, value=pos["venc"])
        c.number_format = 'DD/MM/YYYY'
    c.border = thin_border

for i, w in enumerate([18, 10, 14, 14, 14, 10, 12, 10, 14], 1):
    ws_port.column_dimensions[chr(64 + i)].width = w

# Anchos hoja Opciones
col_widths = [18, 12, 8, 10, 14, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

# Guardar
filename = "ggal_opciones.xlsx"
wb.save(filename)

print(f"Excel generado: {filename}")
print(f"  Hoja 'Opciones': {total_options} opciones GGAL")
print(f"    - {len(OPCIONES_ABRIL)} opciones Abril (29 calls + 29 puts)")
print(f"    - {len(OPCIONES_JUNIO)} opciones Junio (18 calls + 18 puts)")
print(f"  Hoja 'Subyacentes': GGAL acción + 2 futuros Rofex")
print(f"  Hoja 'Portfolio': {len(posiciones)} posiciones ejemplo")
print(f"\n  GGAL Spot: ${GGAL_SPOT}")
print(f"  Futuro Abril: ${round(GGAL_SPOT * 1.035, 2)}")
print(f"  Futuro Junio: ${round(GGAL_SPOT * 1.075, 2)}")
print(f"\nAbrir {filename} en Excel y ejecutar el bridge.")
