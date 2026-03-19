"""
Genera un Excel de ejemplo con datos simulados de opciones argentinas
similar a lo que se vería en etrader v3.75.9.

Uso (en Windows PowerShell):
    python create_sample_excel.py

Genera: etrader_sample.xlsx
"""
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, date, timedelta
import random

wb = Workbook()

# ============================================================
# HOJA 1: Opciones (Option Chain)
# ============================================================
ws_opciones = wb.active
ws_opciones.title = "Opciones"

# Estilos
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
call_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
put_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Headers
headers = [
    "Simbolo", "Subyacente", "Tipo", "Strike", "Vencimiento",
    "Ultimo", "Bid", "Ask", "Volumen", "VI",
    "OI", "Spot", "Delta", "Gamma", "Theta", "Vega"
]
for col, header in enumerate(headers, 1):
    cell = ws_opciones.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Datos de opciones argentinas simulados
subyacentes = {
    "GGAL": {"spot": 5850.0, "vol_base": 0.55},
    "YPF": {"spot": 42500.0, "vol_base": 0.48},
    "PAMP": {"spot": 3200.0, "vol_base": 0.52},
    "COME": {"spot": 380.0, "vol_base": 0.60},
    "TXAR": {"spot": 1150.0, "vol_base": 0.50},
}

# Vencimientos (tercer viernes del mes)
vencimientos = [
    date(2026, 4, 17),   # Abril
    date(2026, 5, 15),   # Mayo
    date(2026, 6, 19),   # Junio
]

row = 2
for underlying, info in subyacentes.items():
    spot = info["spot"]
    vol_base = info["vol_base"]

    for venc in vencimientos:
        # Generar strikes alrededor del spot
        if spot > 10000:
            step = 2500
        elif spot > 1000:
            step = 250
        elif spot > 100:
            step = 50
        else:
            step = 20

        strikes = []
        base = round(spot / step) * step
        for i in range(-3, 4):
            strikes.append(base + i * step)

        month_code = venc.strftime("%m")
        year_code = venc.strftime("%y")

        for strike in strikes:
            for opt_type in ["C", "P"]:
                # Calcular precio simulado
                moneyness = (spot - strike) / spot if opt_type == "C" else (strike - spot) / spot
                time_to_exp = (venc - date.today()).days / 365.0
                if time_to_exp <= 0:
                    continue

                # Precio simplificado
                intrinsic = max(0, spot - strike) if opt_type == "C" else max(0, strike - spot)
                time_value = spot * vol_base * (time_to_exp ** 0.5) * 0.4
                price = intrinsic + time_value * max(0.1, 1 - abs(moneyness) * 3)
                price = max(1.0, round(price, 2))

                bid = round(price * 0.97, 2)
                ask = round(price * 1.03, 2)
                vol = random.randint(10, 5000)
                oi = random.randint(100, 50000)

                # Volatilidad implícita simulada (smile)
                iv = vol_base + abs(moneyness) * 0.15 + random.uniform(-0.03, 0.03)
                iv = round(iv, 4)

                # Greeks simplificados
                if opt_type == "C":
                    delta = round(max(0.01, min(0.99, 0.5 + moneyness * 2)), 4)
                else:
                    delta = round(max(-0.99, min(-0.01, -0.5 + moneyness * 2)), 4)
                gamma = round(0.01 * (1 - abs(moneyness) * 2), 4)
                theta = round(-price * 0.01 / max(0.01, time_to_exp), 4)
                vega = round(spot * 0.01 * (time_to_exp ** 0.5), 4)

                # Símbolo estilo etrader: GFGC5850.AB (Galicia Call 5850 Abril Base)
                tipo_letra = "C" if opt_type == "C" else "V"  # C=Call, V=Venta(Put) en Argentina
                mes_letras = ["", "EN", "FE", "MR", "AB", "MY", "JN", "JL", "AG", "SE", "OC", "NO", "DI"]
                mes = mes_letras[venc.month]
                simbolo = f"GF{tipo_letra}{int(strike)}.{mes}"
                if underlying != "GGAL":
                    prefix = underlying[:2]
                    simbolo = f"{prefix}{tipo_letra}{int(strike)}.{mes}"

                tipo_texto = "CALL" if opt_type == "C" else "PUT"

                ws_opciones.cell(row=row, column=1, value=simbolo).border = thin_border
                ws_opciones.cell(row=row, column=2, value=underlying).border = thin_border
                ws_opciones.cell(row=row, column=3, value=tipo_texto).border = thin_border
                ws_opciones.cell(row=row, column=4, value=strike).border = thin_border
                ws_opciones.cell(row=row, column=5, value=venc).border = thin_border
                ws_opciones.cell(row=row, column=5).number_format = 'DD/MM/YYYY'
                ws_opciones.cell(row=row, column=6, value=price).border = thin_border
                ws_opciones.cell(row=row, column=7, value=bid).border = thin_border
                ws_opciones.cell(row=row, column=8, value=ask).border = thin_border
                ws_opciones.cell(row=row, column=9, value=vol).border = thin_border
                ws_opciones.cell(row=row, column=10, value=iv).border = thin_border
                ws_opciones.cell(row=row, column=10).number_format = '0.00%'
                ws_opciones.cell(row=row, column=11, value=oi).border = thin_border
                ws_opciones.cell(row=row, column=12, value=spot).border = thin_border
                ws_opciones.cell(row=row, column=13, value=delta).border = thin_border
                ws_opciones.cell(row=row, column=14, value=gamma).border = thin_border
                ws_opciones.cell(row=row, column=15, value=theta).border = thin_border
                ws_opciones.cell(row=row, column=16, value=vega).border = thin_border

                # Colorear filas
                fill = call_fill if opt_type == "C" else put_fill
                for c in range(1, 17):
                    ws_opciones.cell(row=row, column=c).fill = fill

                row += 1

# Ajustar anchos
col_widths = [18, 12, 8, 10, 14, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]
for i, w in enumerate(col_widths, 1):
    ws_opciones.column_dimensions[chr(64 + i)].width = w

# ============================================================
# HOJA 2: Portfolio (Posiciones actuales)
# ============================================================
ws_portfolio = wb.create_sheet("Portfolio")

portfolio_headers = [
    "Simbolo", "Cantidad", "Precio Entrada", "Precio Actual",
    "P&L", "Tipo", "Subyacente", "Strike", "Vencimiento"
]
for col, header in enumerate(portfolio_headers, 1):
    cell = ws_portfolio.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Posiciones de ejemplo
posiciones = [
    {"simbolo": "GFC5850.AB", "cant": 10, "entrada": 320.0, "actual": 385.0, "tipo": "CALL", "sub": "GGAL", "strike": 5850, "venc": date(2026, 4, 17)},
    {"simbolo": "GFV6100.AB", "cant": -5, "entrada": 280.0, "actual": 245.0, "tipo": "PUT", "sub": "GGAL", "strike": 6100, "venc": date(2026, 4, 17)},
    {"simbolo": "GFC6350.MY", "cant": 20, "entrada": 150.0, "actual": 210.0, "tipo": "CALL", "sub": "GGAL", "strike": 6350, "venc": date(2026, 5, 15)},
    {"simbolo": "YPC42500.AB", "cant": 5, "entrada": 2800.0, "actual": 3100.0, "tipo": "CALL", "sub": "YPF", "strike": 42500, "venc": date(2026, 4, 17)},
    {"simbolo": "YPV40000.AB", "cant": -3, "entrada": 1500.0, "actual": 1200.0, "tipo": "PUT", "sub": "YPF", "strike": 40000, "venc": date(2026, 4, 17)},
    {"simbolo": "PAC3200.JN", "cant": 15, "entrada": 180.0, "actual": 220.0, "tipo": "CALL", "sub": "PAMP", "strike": 3200, "venc": date(2026, 6, 19)},
    {"simbolo": "GFC5600.JN", "cant": 8, "entrada": 520.0, "actual": 610.0, "tipo": "CALL", "sub": "GGAL", "strike": 5600, "venc": date(2026, 6, 19)},
    {"simbolo": "GFV5350.MY", "cant": -10, "entrada": 95.0, "actual": 70.0, "tipo": "PUT", "sub": "GGAL", "strike": 5350, "venc": date(2026, 5, 15)},
]

for i, pos in enumerate(posiciones, 2):
    pnl = (pos["actual"] - pos["entrada"]) * pos["cant"]

    ws_portfolio.cell(row=i, column=1, value=pos["simbolo"]).border = thin_border
    ws_portfolio.cell(row=i, column=2, value=pos["cant"]).border = thin_border
    ws_portfolio.cell(row=i, column=3, value=pos["entrada"]).border = thin_border
    ws_portfolio.cell(row=i, column=4, value=pos["actual"]).border = thin_border
    cell_pnl = ws_portfolio.cell(row=i, column=5, value=pnl)
    cell_pnl.border = thin_border
    cell_pnl.font = Font(color="006400" if pnl >= 0 else "8B0000", bold=True)
    ws_portfolio.cell(row=i, column=6, value=pos["tipo"]).border = thin_border
    ws_portfolio.cell(row=i, column=7, value=pos["sub"]).border = thin_border
    ws_portfolio.cell(row=i, column=8, value=pos["strike"]).border = thin_border
    ws_portfolio.cell(row=i, column=9, value=pos["venc"]).border = thin_border
    ws_portfolio.cell(row=i, column=9).number_format = 'DD/MM/YYYY'

# Ajustar anchos
for i, w in enumerate([18, 10, 14, 14, 14, 8, 12, 10, 14], 1):
    ws_portfolio.column_dimensions[chr(64 + i)].width = w

# ============================================================
# HOJA 3: Resumen
# ============================================================
ws_resumen = wb.create_sheet("Resumen")
ws_resumen.cell(row=1, column=1, value="Galleguimetro - Datos de Ejemplo").font = Font(bold=True, size=14)
ws_resumen.cell(row=3, column=1, value="Este archivo simula datos de etrader v3.75.9")
ws_resumen.cell(row=4, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
ws_resumen.cell(row=5, column=1, value=f"Opciones: {row - 2} registros")
ws_resumen.cell(row=6, column=1, value=f"Posiciones: {len(posiciones)} registros")
ws_resumen.cell(row=8, column=1, value="Subyacentes:").font = Font(bold=True)
for i, (sym, info) in enumerate(subyacentes.items(), 9):
    ws_resumen.cell(row=i, column=1, value=sym)
    ws_resumen.cell(row=i, column=2, value=f"Spot: ${info['spot']:.2f}")
    ws_resumen.cell(row=i, column=3, value=f"Vol base: {info['vol_base']:.0%}")

ws_resumen.cell(row=15, column=1, value="Instrucciones:").font = Font(bold=True)
ws_resumen.cell(row=16, column=1, value="1. Abrir este archivo en Excel")
ws_resumen.cell(row=17, column=1, value="2. Ejecutar el bridge: python dde_bridge.py --backend-url http://IP:8000 --username admin --password admin12345")
ws_resumen.cell(row=18, column=1, value="3. El bridge leerá la hoja 'Opciones' y 'Portfolio' automáticamente")
ws_resumen.column_dimensions['A'].width = 80

# Guardar
filename = "etrader_sample.xlsx"
wb.save(filename)
print(f"Excel generado: {filename}")
print(f"  - Hoja 'Opciones': {row - 2} opciones de {len(subyacentes)} subyacentes")
print(f"  - Hoja 'Portfolio': {len(posiciones)} posiciones")
print(f"  - Hoja 'Resumen': info general")
print(f"\nAbrir {filename} en Excel y luego ejecutar el bridge.")
